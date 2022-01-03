from openpyxl import load_workbook
wb = load_workbook("sample_merge.xlsx")
ws = wb.active

# B2 부터 D2까지 병합되어 있던 셀을 분리한다
ws.unmerge_cells("B2:D2")
wb.save("sample_unmerge.xlsx")